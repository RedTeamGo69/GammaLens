"""Ticker-keyed forward-test workbook in ui_spread_finder.

The weekly Excel export's Scoreboard used to 3D-sum fixed rows across week
tabs, which silently mis-scored any week whose ticker mix or row order
differed. These pin the row-agnostic rework:

  * the Scoreboard matches instruments BY NAME via SUMIF/COUNTIFS over
    INDIRECT on pre-listed Monday tab names (hidden column M),
  * missing week tabs resolve through the hidden sanitizer column L to the
    empty FTBlank sheet — NOT via IFERROR, which Excel refuses to array-lift
    around SUMIF-over-INDIRECT (verified in real Excel: the whole product
    silently collapses to 0),
  * per-tier hidden denominators only count weeks where the tier actually
    had a band and a filled close,
  * week tabs stay self-contained (no cross-sheet refs → Move/Copy safe),
  * blank Scoreboard slots ship armed so typing a ticker starts tracking it.
"""
from datetime import date, timedelta
from io import BytesIO

import openpyxl
import pytest

import ui_spread_finder as usf

WEEK = "2026-07-06"   # a Monday


def _row(ticker, bands=None, error=None, ref=100.0, prev=99.0, notes=None):
    return {"ticker": ticker, "ref": ref, "prev_close": prev,
            "bands": bands or {}, "notes": notes or [], "error": error}


def _full_bands():
    return {"lower_pi": (95.0, 105.0), "point": (93.0, 107.0),
            "pi_upper": (90.0, 110.0), "effective": (88.0, 112.0)}


@pytest.fixture(scope="module")
def workbook():
    data = usf._build_forward_test_workbook(
        week_start=WEEK, model_choice="M3_extended",
        rows=[_row("SPX", _full_bands()),
              _row("QQQ", _full_bands()),
              _row("AMD", error="data collection failed")],
    )
    return openpyxl.load_workbook(BytesIO(data))


def test_sheets_are_scoreboard_week_tab_and_blank(workbook):
    # Bookends are gone — the Scoreboard finds week tabs by name now, and
    # unpasted weeks resolve to the empty FTBlank utility sheet.
    assert workbook.sheetnames == ["Scoreboard", WEEK, usf._FT_BLANK_SHEET]


def test_blank_sheet_scan_region_is_empty(workbook):
    bk = workbook[usf._FT_BLANK_SHEET]
    for r in range(usf._FT_FIRST_DATA_ROW, usf._FT_LAST_DATA_ROW + 1):
        for col in "BOPQRS":
            assert bk[f"{col}{r}"].value is None


def test_sanitizer_maps_missing_tabs_to_blank_sheet(workbook):
    sb = workbook["Scoreboard"]
    first = usf._FT_SB_FIRST_ROW
    n = len(usf._ft_monday_tab_names(WEEK))
    for k in (0, n - 1):                       # first and last listed Monday
        r = first + k
        assert sb[f"L{r}"].value == (
            f'=IF(ISREF(INDIRECT("\'"&$M{r}&"\'!$A$1")),'
            f'$M{r},"{usf._FT_BLANK_SHEET}")'
        )
    assert sb.column_dimensions["L"].hidden


def test_monday_tab_names_cover_back_and_forward_horizon():
    names = usf._ft_monday_tab_names(WEEK)
    assert len(names) == usf._FT_SB_WEEKS_BACK + usf._FT_SB_WEEKS_FWD + 1
    assert names[usf._FT_SB_WEEKS_BACK] == WEEK
    mondays = [date.fromisoformat(n) for n in names]
    assert all(d.weekday() == 0 for d in mondays)
    assert all((b - a) == timedelta(weeks=1) for a, b in zip(mondays, mondays[1:]))


def test_scoreboard_prelists_names_as_text(workbook):
    sb = workbook["Scoreboard"]
    first = usf._FT_SB_FIRST_ROW
    names = usf._ft_monday_tab_names(WEEK)
    got = [sb[f"M{first + k}"].value for k in range(len(names))]
    assert got == names
    assert sb[f"M{first}"].number_format == "@"       # never a date serial
    assert sb.column_dimensions["M"].hidden


def test_scoreboard_scores_by_ticker_not_row(workbook):
    sb = workbook["Scoreboard"]
    r = usf._FT_SB_FIRST_ROW
    assert sb[f"A{r}"].value == "SPX"
    weeks = sb[f"B{r}"].value
    # Weeks Scored: SUMIF on the Instrument column (B) of every listed tab,
    # summing the Scored? flags (S) — keyed on $A, no fixed row anywhere,
    # through the sanitized names in L (never IFERROR — see module docstring).
    assert "SUMIF(" in weeks and "INDIRECT(" in weeks and f"$A{r}" in weeks
    assert "$B$6:$B$45" in weeks and "$S$6:$S$45" in weeks
    assert '"\'"&$L$6:' in weeks
    assert "IFERROR" not in weeks
    assert "WeeksStart" not in weeks and "WeeksEnd" not in weeks
    # Wins count flag=1 matches for that ticker.
    wins = sb[f"C{r}"].value
    assert "COUNTIFS(" in wins and wins.endswith(",1)))")
    # Hidden per-tier denominator counts numeric flags only (">=0") AND
    # requires the tier's Low band to hold a number (">0"): pre-rework tabs'
    # unguarded flags emit 0 on band-less rows — the band criterion keeps
    # those from counting as losses.
    denom = sb[f"N{r}"].value
    assert '">=0"' in denom and '">0"' in denom
    assert sb.column_dimensions["N"].hidden
    # Hit % divides wins by its own tier denominator, not Weeks Scored.
    assert f"C{r}/N{r}" in sb[f"D{r}"].value
    assert sb[f"D{r}"].number_format == "0%"


def test_scoreboard_tier_column_pairings(workbook):
    # Pin the (win, denominator, flag, band-Low) mapping for ALL four tiers —
    # a transposition in any tier must fail loudly, not just tier 1.
    sb = workbook["Scoreboard"]
    r = usf._FT_SB_FIRST_ROW
    pairings = [("C", "N", "O", "G"), ("E", "O", "P", "I"),
                ("G", "P", "Q", "K"), ("I", "Q", "R", "M")]
    for hit_col, (win_col, den_col, flag_col, lo_col) in zip("DFHJ", pairings):
        flag_rng = f"${flag_col}$6:${flag_col}$45"
        lo_rng = f"${lo_col}$6:${lo_col}$45"
        assert flag_rng in sb[f"{win_col}{r}"].value
        assert flag_rng in sb[f"{den_col}{r}"].value
        assert lo_rng in sb[f"{den_col}{r}"].value
        assert f"{win_col}{r}/{den_col}{r}" in sb[f"{hit_col}{r}"].value


def test_scoreboard_diagnostics_and_dup_detector(workbook):
    sb = workbook["Scoreboard"]
    k1 = sb["K1"].value
    # Live tab counter + the two loud warnings for silent failure modes.
    assert k1.startswith('="Week tabs found: "')
    assert usf._FT_BLANK_SHEET in k1
    assert "(2)" in k1 and "TODAY()" in k1
    # Hidden R column flags a duplicate tab Excel renamed to "<Monday> (2)".
    r = usf._FT_SB_FIRST_ROW
    assert sb[f"R{r}"].value == (
        f'=IF(ISREF(INDIRECT("\'"&$M{r}&" (2)\'!$A$1")),1,0)'
    )
    assert sb.column_dimensions["R"].hidden


def test_scoreboard_protected_with_editable_slots(workbook):
    sb = workbook["Scoreboard"]
    assert sb.protection.sheet
    # Instrument slots (filled AND blank) must accept typing under protection;
    # everything else — formulas, hidden plumbing — stays locked.
    for r in (usf._FT_SB_FIRST_ROW, usf._FT_SB_LAST_ROW):
        assert sb[f"A{r}"].protection.locked is False
        assert sb[f"B{r}"].protection.locked is not False
    assert sb[f"L{usf._FT_SB_FIRST_ROW}"].protection.locked is not False


def test_blank_scoreboard_slots_ship_armed(workbook):
    sb = workbook["Scoreboard"]
    r = usf._FT_SB_LAST_ROW          # far below the 3 exported tickers
    assert sb[f"A{r}"].value is None
    for col in "BCDEFGHIJ":
        assert str(sb[f"{col}{r}"].value).startswith(f'=IF($A{r}="",""')


def test_scoreboard_all_row_pools_all_slots(workbook):
    sb = workbook["Scoreboard"]
    ar, lo, hi = usf._FT_SB_ALL_ROW, usf._FT_SB_FIRST_ROW, usf._FT_SB_LAST_ROW
    assert sb[f"A{ar}"].value == "ALL"
    assert sb[f"B{ar}"].value == f"=SUM(B{lo}:B{hi})"
    assert sb[f"D{ar}"].value == f'=IF(N{ar}=0,"—",C{ar}/N{ar})'


def test_week_tab_flags_guard_missing_bands_and_text(workbook):
    ws = workbook[WEEK]
    r = usf._FT_FIRST_DATA_ROW
    # COUNT counts numbers only, so the guard covers BOTH failure modes: a
    # band-less tier stays blank (unguarded compare against empty cells gives
    # a false 0) and a TEXT close/band stays blank (any text compares above
    # any number in Excel, so it too would score a false loss).
    assert ws[f"O{r}"].value == (
        f'=IF(OR($B{r}="",COUNT($E{r},G{r},H{r})<3),"",'
        f'IF(AND($E{r}>=G{r},$E{r}<=H{r}),1,0))'
    )
    # Scored? = instrument present, numeric close, ≥1 numeric band.
    assert ws[f"S{r}"].value == (
        f'=IF(OR($B{r}="",COUNT($E{r})=0,COUNT($G{r}:$N{r})=0),"",1)'
    )


def test_week_tab_rows_ship_armed_to_scan_edge(workbook):
    # Rows beyond the exported tickers carry the full formula set (guarded on
    # column B) so a hand-typed instrument row scores like an exported one,
    # and the conditional formatting colours the whole armed window.
    ws = workbook[WEEK]
    r = usf._FT_LAST_DATA_ROW                  # far below the 3 exported rows
    assert ws[f"B{r}"].value is None
    for flag_col in "OPQR":
        assert str(ws[f"{flag_col}{r}"].value).startswith(f'=IF(OR($B{r}=""')
    assert str(ws[f"S{r}"].value).startswith(f'=IF(OR($B{r}=""')
    cf_ranges = {str(rng) for cf in ws.conditional_formatting for rng in cf.sqref.ranges}
    assert f"O{usf._FT_FIRST_DATA_ROW}:R{usf._FT_LAST_DATA_ROW}" in cf_ranges


def test_week_tab_protected_with_editable_data_cells(workbook):
    ws = workbook[WEEK]
    assert ws.protection.sheet
    r = usf._FT_FIRST_DATA_ROW
    # Data entry (ticker, ref, close, bands, notes) stays editable — enough
    # to hand-add a full instrument row; flags/Scored? and structure locked.
    for col in "BCDEFGHIJKLMNT":
        assert ws[f"{col}{r}"].protection.locked is False
    for col in "OPQRS":
        assert ws[f"{col}{r}"].protection.locked is not False


def test_week_close_column_validates_numeric(workbook):
    ws = workbook[WEEK]
    dvs = list(ws.data_validations.dataValidation)
    assert any(
        dv.type == "decimal"
        and f"E{usf._FT_FIRST_DATA_ROW}:E{usf._FT_LAST_DATA_ROW}" in str(dv.sqref)
        for dv in dvs
    )


def test_blank_sheet_is_protected(workbook):
    # A stray row typed into FTBlank would be counted once per missing week.
    assert workbook[usf._FT_BLANK_SHEET].protection.sheet


def test_week_tab_is_self_contained(workbook):
    # No cross-sheet references — Move/Copy into the master must never
    # create an external link back to the downloaded file.
    ws = workbook[WEEK]
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                assert "!" not in v, f"{cell.coordinate} references another sheet: {v}"


def test_error_row_carries_note_and_no_bands(workbook):
    ws = workbook[WEEK]
    r = usf._FT_FIRST_DATA_ROW + 2   # AMD
    assert ws[f"B{r}"].value == "AMD"
    assert "data collection failed" in ws[f"T{r}"].value
    assert all(ws[f"{c}{r}"].value is None for c in "GHIJKLMN")


def test_week_rows_capped_at_scan_window():
    many = [_row(f"T{i:02d}", _full_bands()) for i in range(usf._FT_MAX_TICKERS + 5)]
    data = usf._build_forward_test_workbook(
        week_start=WEEK, model_choice="M2_vix", rows=many)
    wb = openpyxl.load_workbook(BytesIO(data))
    ws = wb[WEEK]
    assert ws[f"B{usf._FT_LAST_DATA_ROW}"].value == f"T{usf._FT_MAX_TICKERS - 1:02d}"
    assert ws[f"B{usf._FT_LAST_DATA_ROW + 1}"].value is None
